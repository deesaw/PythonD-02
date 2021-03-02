from openpyxl import Workbook

wb = Workbook()
ws1 = wb.create_sheet("Hello1")
ws2 = wb.create_sheet("Hello2")
ws3 = wb.create_sheet("Hello3")

wb.save("test.xlsx")