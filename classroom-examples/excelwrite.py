from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Deloitee"

ws.cell(row=1, column=1, value="Hello")
ws.cell(row=1, column=2, value="Deloitee")

ws.cell(row=4, column=1, value=456)
ws.cell(row=4, column=2, value=908)

for row in range(5,15):
	for column in range(5,15):
		ws.cell(row=row, column=column, value=row*column)
wb.save("sample.xlsx")

