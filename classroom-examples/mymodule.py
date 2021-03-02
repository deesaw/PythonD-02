import os
import time
import sqlite3
from openpyxl import load_workbook, Workbook

def dump_table_to_xlsx(conn,table,xlfile):
	wb = Workbook()
	ws = wb.active
	cursor = conn.cursor()
	cursor.execute(f"SELECT * from {table}")
	for row in cursor:
		ws.append(row)
	wb.save(xlfile)

def get_ips_from_db(database,tablename):
	conn = sqlite3.connect(database)
	cursor = conn.cursor()
	SQL = """
	SELECT ip FROM {}
	""".format(tablename)
	cursor.execute(SQL)
	ips = [ip[0] for ip in cursor.fetchall()]
	conn.close()
	return ips

def get_ips_from_xlsx(xlfile):
	return [cell.value for cell in load_workbook("ips.xlsx").active['A']]

def isup(ip):
	return True

def get_ips_from_file(filename):
	f=open(filename,"r")
	ips=[]
	for line in f:
		ips.append(line.strip())
	f.close()
	return ips

class FileInfo:
	"""
	This class helps to you get information of a file.s
	"""
	def __init__(self,filename):
		self.filename=filename
		self.f = open(self.filename,'r')
		self.lines=self.f.readlines()
	def getsize(self):
		print(os.path.getsize(self.filename), "bytes")
	def linecount(self):
		print("line count = ",len(self.lines))
	def firstline(self):
		self.nthline(1)
	def lastline(self):
		self.nthline(0)
	def nthline(self,lineno):
		print(self.lines[lineno - 1].rstrip())
	def search(self,pattern):
		"""
		search will help you to search for a pattern
		case insensitive in the file.
		syntax is f.search("pattern")
		"""
		for lineno,line in enumerate(self.lines):
			if pattern.lower() in line.lower():
				print(lineno + 1,line.rstrip())
	def getage(self):
		print("File created on ",time.ctime(os.path.getctime(self.filename)))
		filectime = os.path.getctime(self.filename)
		now = time.time()
		age = now - filectime
		print("Age = {} seconds".format(round(age)))
		print(f"Age = {round(age)} seconds")
	def close(self):
		self.f.close()
		print("File Closed")

headache = Exception("Too much python")
