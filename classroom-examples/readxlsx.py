from openpyxl import load_workbook
wb = load_workbook("marks.xlsx")
#print sheetnames
print(wb.sheetnames)
#select a sheet
ws = wb["Marks"]
#print max rows and max columns
print(ws.max_row,ws.max_column)

#read a cell
print(ws["A3"].value)
print(ws.cell(row = 3, column = 1).value)

#read a row
myrow = ws[1]
for mycell in myrow:
	print(mycell.value, end =" ")
print()
#read a column
print(dir(ws))
mycol = ws['A']
for mycell in mycol:
	print(mycell.value)
#read a range
myrange = ws['A1:C3']
for myrow in myrange:
	for mycell in myrow:
		print(mycell.value, end =" ")
	print()

#generic read

for row in ws.values:
	print(row)